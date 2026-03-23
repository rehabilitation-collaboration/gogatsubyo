"""
統計局 人口推計 年齢各歳別ExcelファイルをCSVに変換
単位: 千人

出力: population_by_age_group_2009_2024.csv
  columns: year, age_group, total_1000, male_1000, female_1000
  age_group: 0-19, 20-29, 30-39, 40-49, 50-59, 60-69, 70-79, 80plus, total
"""

import os
import csv
import re
import io

SAVE_DIR = os.path.dirname(os.path.abspath(__file__))

AGE_GROUPS = {
    "0-19":   list(range(0, 20)),
    "20-29":  list(range(20, 30)),
    "30-39":  list(range(30, 40)),
    "40-49":  list(range(40, 50)),
    "50-59":  list(range(50, 60)),
    "60-69":  list(range(60, 70)),
    "70-79":  list(range(70, 80)),
    "80plus": list(range(80, 120)),
}


def extract_age(label):
    """年齢ラベルから整数を返す。総数はTOTAL_SENTINEL、不明はNone"""
    if label is None:
        return None
    s = str(label).strip()
    if re.match(r"^総\s*数$", s):
        return "total"
    if re.search(r"100\s*歳以上|100 and over|100\+", s):
        return 100
    m = re.search(r"\d+", s)
    if m:
        return int(m.group())
    return None


def _get_float(val):
    if val in (None, "", "-", "－"):
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def read_format_single_col(rows):
    """
    一列形式（2010, 2015, 2020年）:
    col1=年齢ラベル, col3=総数計, col4=男, col5=女
    """
    result = {}
    for row in rows:
        if len(row) < 6:
            continue
        label = row[1]  # col index 1
        age = extract_age(label)
        if age is None:
            continue
        tot = _get_float(row[3])
        mal = _get_float(row[4])
        fem = _get_float(row[5])
        if tot > 0 or age == "total":
            result[age] = (tot, mal, fem)
    return result


def read_format_two_col(rows):
    """
    二列形式（2009, 2011-2024年）:
    左: col0=年齢, col1=計, col2=男, col3=女
    右: col9=年齢, col10=計, col11=男, col12=女
    """
    result = {}
    for row in rows:
        # 左側
        if len(row) > 0:
            age_l = extract_age(row[0])
            if age_l is not None:
                tot = _get_float(row[1] if len(row) > 1 else None)
                mal = _get_float(row[2] if len(row) > 2 else None)
                fem = _get_float(row[3] if len(row) > 3 else None)
                if tot > 0 or age_l == "total":
                    result[age_l] = (tot, mal, fem)

        # 右側
        if len(row) > 9:
            age_r = extract_age(row[9])
            if age_r is not None and age_r != "total":
                tot = _get_float(row[10] if len(row) > 10 else None)
                mal = _get_float(row[11] if len(row) > 11 else None)
                fem = _get_float(row[12] if len(row) > 12 else None)
                if tot > 0:
                    result[age_r] = (tot, mal, fem)
    return result


def detect_format(rows):
    """ファイルのフォーマットを判定する"""
    for row in rows[:20]:
        # 一列形式のシグネチャ: col9に '総  人  口' と英語が共存
        if len(row) > 3 and row[1] is not None:
            s = str(row[1])
            if "年      齢" in s or "Age" in s:
                # 次の行で値列を確認
                pass
        if len(row) > 3:
            if str(row[3] or "").strip() in ("Both sexes", "男 女 計"):
                return "single_col"
    return "two_col"


def read_excel_sheet(path):
    ext = os.path.splitext(path)[1].lower()
    rows = []

    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
    else:
        # .xls: まずBytesIOでopenpyxl試行（PKマジック）、失敗したらxlrd
        with open(path, "rb") as f:
            data = f.read()

        if data[:2] == b"PK":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
        else:
            import xlrd
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            for i in range(ws.nrows):
                rows.append(ws.row_values(i))

    fmt = detect_format(rows)
    if fmt == "single_col":
        return read_format_single_col(rows)
    else:
        return read_format_two_col(rows)


def aggregate_age_groups(age_data):
    records = []

    # 総数（直接取得）
    if "total" in age_data:
        t, m, f = age_data["total"]
        records.append(("total", t, m, f))

    # 年齢グループ別
    for group_name, ages in AGE_GROUPS.items():
        t_sum = m_sum = f_sum = 0.0
        for age in ages:
            if age in age_data:
                t, m, f = age_data[age]
                t_sum += t
                m_sum += m
                f_sum += f
        records.append((group_name, t_sum, m_sum, f_sum))

    return records


def main():
    all_rows = []

    years = sorted([
        int(re.search(r"\d+", f).group())
        for f in os.listdir(SAVE_DIR)
        if f.startswith("pop") and (f.endswith(".xls") or f.endswith(".xlsx"))
    ])

    print(f"処理対象: {years}")

    for year in years:
        fname = None
        for ext in [".xlsx", ".xls"]:
            candidate = os.path.join(SAVE_DIR, f"pop{year}{ext}")
            if os.path.exists(candidate):
                fname = candidate
                break

        if fname is None:
            print(f"[SKIP] {year}: ファイルなし")
            continue

        try:
            age_data = read_excel_sheet(fname)
            records = aggregate_age_groups(age_data)
            for (group, tot, mal, fem) in records:
                all_rows.append({
                    "year": year,
                    "age_group": group,
                    "total_1000": round(tot, 1),
                    "male_1000": round(mal, 1),
                    "female_1000": round(fem, 1),
                })
            n_ages = sum(1 for k in age_data if k != "total")
            print(f"[OK] {year}: {n_ages} 歳分, format={'single' if 'total' in age_data and age_data.get('total',()) else 'unknown'}")
        except Exception as e:
            print(f"[FAIL] {year}: {e}")
            import traceback; traceback.print_exc()

    out_path = os.path.join(SAVE_DIR, "population_by_age_group_2009_2024.csv")
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["year", "age_group", "total_1000", "male_1000", "female_1000"])
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\n出力完了: {out_path}")
    print(f"行数: {len(all_rows)}")

    totals = {r["year"]: r for r in all_rows if r["age_group"] == "total"}
    print("\n年別総数（千人）:")
    for y in sorted(totals.keys()):
        r = totals[y]
        print(f"  {y}: {r['total_1000']:,.0f}千人 (男{r['male_1000']:,.0f} 女{r['female_1000']:,.0f})")

    expected_years = set(range(2009, 2025))
    missing = expected_years - set(totals.keys())
    if missing:
        print(f"\n警告: 欠損年 {sorted(missing)}")


if __name__ == "__main__":
    main()
